"""
Быстрое формирование заказа из Excel файла.
"""
from aiogram import Router, F
from aiogram.types import Message, BufferedInputFile, CallbackQuery, InlineKeyboardMarkup, InlineKeyboardButton, ReplyKeyboardMarkup, KeyboardButton, ReplyKeyboardRemove
from aiogram.fsm.context import FSMContext
from aiogram.utils.keyboard import InlineKeyboardBuilder, ReplyKeyboardBuilder
from core.parser import ExcelParser
from openpyxl import load_workbook
from io import BytesIO
from datetime import datetime
from pathlib import Path
from typing import List, Dict
import json
import re
import pandas as pd
import os
import hashlib
from database.crud import async_session_maker, get_or_create_user, create_quick_order
from bot.states import QuickOrderStates

router = Router()

# Простой кэш для результатов парсинга
_parsing_cache = {}

def get_file_hash(file_path: str) -> str:
    """Получить хэш файла для кэширования."""
    with open(file_path, 'rb') as f:
        return hashlib.md5(f.read()).hexdigest()

def get_cached_parsing_result(file_path: str) -> List[Dict]:
    """Получить кэшированный результат парсинга."""
    file_hash = get_file_hash(file_path)
    return _parsing_cache.get(file_hash)

def cache_parsing_result(file_path: str, result: List[Dict]):
    """Сохранить результат парсинга в кэш."""
    file_hash = get_file_hash(file_path)
    _parsing_cache[file_hash] = result
    # Ограничиваем размер кэша (максимум 10 файлов)
    if len(_parsing_cache) > 10:
        # Удаляем самый старый элемент
        oldest_key = next(iter(_parsing_cache))
        del _parsing_cache[oldest_key]


@router.message(F.document)
async def process_excel_file(message: Message, state: FSMContext):
    """
    Обработать загруженный Excel файл.
    Парсим, показываем позиции, ждем указания количества.
    """
    document = message.document
    
    # Проверка типа файла
    if not document.file_name.endswith(('.xlsx', '.xls')):
        await message.answer(
            "Пожалуйста, отправьте файл в формате Excel (.xlsx или .xls)."
        )
        return
    
    await message.answer("Парсинг файла...")
    
    # Скачиваем файл
    file_path = f"temp_files/{document.file_name}"
    await message.bot.download(document, destination=file_path)
    
    # Парсинг файла (с кэшированием)
    cached_result = get_cached_parsing_result(file_path)
    if cached_result:
        beer_items = cached_result
        await message.answer(f"Файл загружен из кэша! Найдено {len(beer_items)} позиций")
    else:
        parser = ExcelParser()
        beer_items = parser.parse_file(file_path)
        cache_parsing_result(file_path, beer_items)
    
    if not beer_items:
        await message.answer("Не удалось извлечь данные из файла.")
        return
    
    # Сохраняем данные в состояние
    await state.update_data(
        file_path=file_path,
        filename=document.file_name,
        items=beer_items,
        current_page=0
    )
    
    # Показываем позиции
    sent_message = await show_items_page(message, beer_items, 0, state=state)
    
    # Сохраняем ID сообщения для редактирования
    await state.update_data(list_message_id=sent_message.message_id)
    await state.set_state(QuickOrderStates.viewing_page)


async def show_items_page(message: Message, items: List[Dict], page: int, items_per_page: int = 20, brewery_filter: str = None, edit_message_id: int = None, state: FSMContext = None):
    """Показать страницу с позициями."""
    # Фильтруем по пивоварне если задан фильтр
    if brewery_filter:
        filtered_items = [item for item in items if item.get('пивоварня') == brewery_filter]
    else:
        filtered_items = items
    
    total_items = len(filtered_items)
    total_pages = (total_items + items_per_page - 1) // items_per_page if total_items > 0 else 1
    
    if page >= total_pages:
        page = total_pages - 1
    if page < 0:
        page = 0
    
    start_idx = page * items_per_page
    end_idx = min(start_idx + items_per_page, total_items)
    page_items = filtered_items[start_idx:end_idx]
    
    # Подсчет выбранных позиций
    selected_items = [item for item in items if (item.get('заказ') or 0) > 0]
    selected_count = len(selected_items)
    total_qty_cans = sum((item.get('заказ') or 0) for item in selected_items if 'кег' not in str(item.get('объем', '')).lower())
    total_qty_kegs = sum((item.get('заказ') or 0) for item in selected_items if 'кег' in str(item.get('объем', '')).lower())
    
    # Формируем текст
    text = f"**Найдено позиций: {len(items)}**"
    if brewery_filter:
        text += f" | Фильтр: {brewery_filter}"
    text += "\n\n"
    text += f"Страница {page + 1} из {total_pages} (позиции {start_idx + 1}-{end_idx})\n\n"
    
    # Группируем по пивоварням и сохраняем глобальные индексы
    breweries = {}
    page_items_with_global_idx = []  # (global_idx, item)
    
    for item in page_items:
        # Находим глобальный индекс в полном списке items
        global_idx = None
        for i, full_item in enumerate(items):
            if (full_item.get('название') == item.get('название') and 
                full_item.get('объем') == item.get('объем') and
                full_item.get('_row_index') == item.get('_row_index')):
                global_idx = i + 1
                break
        
        if global_idx:
            page_items_with_global_idx.append((global_idx, item))
            brewery = item.get('пивоварня', 'Без пивоварни')
            if brewery not in breweries:
                breweries[brewery] = []
            breweries[brewery].append((global_idx, item))
    
    for brewery, items_list in breweries.items():
        text += f"**{brewery}**\n"
        
        # Разделяем на кеги и банки
        kegs = []
        cans_bottles = []
        
        for idx, item in items_list:
            volume_lower = str(item.get('объем', '')).lower()
            if 'кег' in volume_lower or 'keg' in volume_lower:
                kegs.append((idx, item))
            else:
                cans_bottles.append((idx, item))
        
        # Сначала показываем кеги
        if kegs:
            text += "\n**КЕГИ:**\n"
            for idx, item in kegs:
                name = item['название']
                if len(name) > 35:
                    display_name = name[:35] + "..."
                else:
                    display_name = name
                
                qty = item.get('заказ') or 0
                checkbox = "✓" if qty > 0 else " "
                volume = item.get('объем', '')
                price = item.get('цена', '')
                stock = item.get('остаток', '')
                
                text += f"`{idx:3d}` [{checkbox}] {display_name}"
                if qty > 0:
                    text += f" **x{qty}**"
                text += "\n"
                
                stock_text = f"      {volume} | {price}"
                if stock:
                    if isinstance(stock, int):
                        stock_text += f" | Остаток: {stock} шт"
                    else:
                        stock_text += f" | {stock}"
                text += stock_text + "\n"
        
        # Потом банки и бутылки
        if cans_bottles:
            text += "\n**БАНКИ/БУТЫЛКИ:**\n"
            for idx, item in cans_bottles:
                name = item['название']
                if len(name) > 35:
                    display_name = name[:35] + "..."
                else:
                    display_name = name
                
                qty = item.get('заказ') or 0
                checkbox = "✓" if qty > 0 else " "
                volume = item.get('объем', '')
                price = item.get('цена', '')
                stock = item.get('остаток', '')
                
                text += f"`{idx:3d}` [{checkbox}] {display_name}"
                if qty > 0:
                    text += f" **x{qty}**"
                text += "\n"
                
                stock_text = f"      {volume} | {price}"
                if stock:
                    if isinstance(stock, int):
                        stock_text += f" | Остаток: {stock} шт"
                    else:
                        stock_text += f" | {stock}"
                text += stock_text + "\n"
        
        text += "\n"
    
    # Итоговая статистика
    if selected_count > 0:
        text += "\n**Выбрано всего:** "
        parts = []
        if total_qty_cans > 0:
            parts.append(f"{total_qty_cans} банок")
        if total_qty_kegs > 0:
            parts.append(f"{total_qty_kegs} кег")
        text += f"{selected_count} позиций ({', '.join(parts)})\n"
    
    text += "\n**Выбор:**\n"
    text += "Введите номер позиции для выбора количества\n"
    text += "Или используйте формат: `номер:кол-во` (например: `1:12`)"
    
    # Получаем уникальные пивоварни для фильтра (с кэшированием)
    all_breweries = None
    if state:
        data = await state.get_data()
        all_breweries = data.get('cached_breweries')
    
    if not all_breweries:
        all_breweries = list(set(item.get('пивоварня', 'Без пивоварни') for item in items))
        if state:
            await state.update_data(cached_breweries=all_breweries)
    
    # Кнопки пагинации + быстрый выбор
    keyboard = get_pagination_keyboard(page, total_pages, "page", selected_count, all_breweries, brewery_filter, page_items_with_global_idx)
    
    # Редактируем существующее сообщение или отправляем новое
    if edit_message_id:
        try:
            await message.bot.edit_message_text(
                text=text,
                chat_id=message.chat.id,
                message_id=edit_message_id,
                parse_mode="Markdown",
                reply_markup=keyboard
            )
            return None  # Сообщение отредактировано
        except:
            pass  # Если не удалось отредактировать, отправим новое
    
    # Отправляем новое сообщение
    return await message.answer(text, parse_mode="Markdown", reply_markup=keyboard)


def get_pagination_keyboard(current_page: int, total_pages: int, prefix: str, selected_count: int = 0, 
                           breweries: List[str] = None, current_brewery: str = None, 
                           page_items_with_idx: List = None) -> InlineKeyboardMarkup:
    """Создать улучшенную клавиатуру навигации."""
    builder = InlineKeyboardBuilder()
    
    # Первый ряд: быстрая навигация
    if total_pages > 1:
        # Кнопка "В начало"
        if current_page > 0:
            builder.add(InlineKeyboardButton(
                text="⏪",
                callback_data=f"{prefix}:0"
            ))
        
        # Кнопка "Назад"
        if current_page > 0:
            builder.add(InlineKeyboardButton(
                text="◀️",
                callback_data=f"{prefix}:{current_page - 1}"
            ))
        
        # Индикатор страницы
        builder.add(InlineKeyboardButton(
            text=f"{current_page + 1}/{total_pages}",
            callback_data="page_info"
        ))
        
        # Кнопка "Вперед"
        if current_page < total_pages - 1:
            builder.add(InlineKeyboardButton(
                text="▶️",
                callback_data=f"{prefix}:{current_page + 1}"
            ))
        
        # Кнопка "В конец"
        if current_page < total_pages - 1:
            builder.add(InlineKeyboardButton(
                text="⏩",
                callback_data=f"{prefix}:{total_pages - 1}"
            ))
    
    # Второй ряд: фильтры и корзина
    row2 = []
    
    # Кнопка корзины
    cart_text = f"Корзина ({selected_count})" if selected_count > 0 else "Корзина"
    row2.append(InlineKeyboardButton(
        text=cart_text,
        callback_data="show_cart"
    ))
    
    # Кнопка поиска
    row2.append(InlineKeyboardButton(
        text="Поиск",
        callback_data="start_search"
    ))
    
    # Кнопка фильтра по пивоварням (если их не слишком много)
    if breweries and len(breweries) <= 10:
        row2.append(InlineKeyboardButton(
            text="Пивоварни",
            callback_data="show_breweries"
        ))
    
    for btn in row2:
        builder.add(btn)
    
    # Третий ряд: сброс фильтра (если активен)
    if current_brewery:
        builder.row(InlineKeyboardButton(
            text="Показать все",
            callback_data="clear_filter"
        ))
    
    # Кнопки быстрого выбора убраны - теперь только через ввод номера позиции
    
    # Ряд завершения заказа
    builder.row(InlineKeyboardButton(
        text="Завершить заказ",
        callback_data="finish_order"
    ))
    
    return builder.as_markup()


def get_quantity_keyboard(is_keg: bool) -> ReplyKeyboardMarkup:
    """Создать клавиатуру выбора количества."""
    builder = ReplyKeyboardBuilder()
    
    if is_keg:
        # Для кег: [1] [2] [3] [5] [10]
        builder.row(
            KeyboardButton(text="1"),
            KeyboardButton(text="2"),
            KeyboardButton(text="3"),
            KeyboardButton(text="5"),
            KeyboardButton(text="10")
        )
    else:
        # Для банок: [10] [12] [20]
        builder.row(
            KeyboardButton(text="10"),
            KeyboardButton(text="12"),
            KeyboardButton(text="20")
        )
    
    # Кнопка отмены
    builder.row(KeyboardButton(text="Отмена"))
    
    return builder.as_markup(resize_keyboard=True, one_time_keyboard=True)


@router.message(QuickOrderStates.viewing_page)
async def handle_position_selection(message: Message, state: FSMContext):
    """Обработка выбора позиции для заказа."""
    if not message.text:
        return
    
    text = message.text.strip()
    
    # Проверяем формат "номер:количество"
    if ":" in text:
        try:
            parts = text.split(":", 1)
            item_idx = int(parts[0].strip())
            qty = int(parts[1].strip())
            
            data = await state.get_data()
            items = data.get('items', [])
            
            if 1 <= item_idx <= len(items):
                # Обновляем количество
                items[item_idx - 1]['заказ'] = qty
                await state.update_data(items=items)
                
                item_name = items[item_idx - 1]['название']
                short_name = item_name[:25] + "..." if len(item_name) > 25 else item_name
                
                await message.answer(f"Добавлено: {short_name} x {qty} шт")
                
                # Обновляем список
                current_page = data.get('current_page', 0)
                brewery_filter = data.get('brewery_filter', None)
                list_message_id = data.get('list_message_id')
                await show_items_page(message, items, current_page, brewery_filter=brewery_filter, edit_message_id=list_message_id, state=state)
            else:
                await message.answer("Ошибка: позиция не найдена")
        except ValueError:
            await message.answer("Неверный формат. Используйте: `номер:количество` (например: `1:12`)")
        return
    
    # Проверяем, является ли текст просто номером позиции
    try:
        item_idx = int(text)
        data = await state.get_data()
        items = data.get('items', [])
        
        if 1 <= item_idx <= len(items):
            item = items[item_idx - 1]
            item_name = item['название']
            
            # Определяем тип позиции (кега или банка)
            is_keg = 'кег' in str(item.get('объем', '')).lower()
            
            # Сохраняем выбранную позицию
            await state.update_data(selected_item_idx=item_idx)
            
            # Показываем клавиатуру выбора количества
            keyboard = get_quantity_keyboard(is_keg)
            
            type_text = "кег" if is_keg else "банок"
            await message.answer(
                f"Выберите количество для позиции {item_idx}:\n"
                f"**{item_name}**\n"
                f"Тип: {type_text}",
                parse_mode="Markdown",
                reply_markup=keyboard
            )
            
            await state.set_state(QuickOrderStates.selecting_quantity)
        else:
            await message.answer("Ошибка: позиция не найдена")
    except ValueError:
        await message.answer("Введите номер позиции или `номер:количество`")


@router.message(QuickOrderStates.selecting_quantity)
async def handle_quantity_selection(message: Message, state: FSMContext):
    """Обработка выбора количества."""
    if not message.text:
        return
        
    text = message.text.strip()
    
    if text == "Отмена":
        await message.answer("Выбор отменен", reply_markup=ReplyKeyboardRemove())
        await state.set_state(QuickOrderStates.viewing_page)
        return
    
    # Проверяем, является ли текст числом
    try:
        qty = int(text)
    except ValueError:
        await message.answer("Введите число или нажмите 'Отмена'")
        return
    
    data = await state.get_data()
    items = data.get('items', [])
    selected_item_idx = data.get('selected_item_idx')
    
    if selected_item_idx and 1 <= selected_item_idx <= len(items):
        # Обновляем количество
        items[selected_item_idx - 1]['заказ'] = qty
        await state.update_data(items=items)
        
        item_name = items[selected_item_idx - 1]['название']
        short_name = item_name[:25] + "..." if len(item_name) > 25 else item_name
        
        await message.answer(f"Добавлено: {short_name} x {qty} шт", reply_markup=ReplyKeyboardRemove())
        
        # Обновляем список
        current_page = data.get('current_page', 0)
        brewery_filter = data.get('brewery_filter', None)
        list_message_id = data.get('list_message_id')
        await show_items_page(message, items, current_page, brewery_filter=brewery_filter, edit_message_id=list_message_id, state=state)
        
        await state.set_state(QuickOrderStates.viewing_page)
    else:
        await message.answer("Ошибка: позиция не найдена", reply_markup=ReplyKeyboardRemove())
        await state.set_state(QuickOrderStates.viewing_page)


@router.callback_query(F.data.startswith("page:"))
async def handle_pagination(callback: CallbackQuery, state: FSMContext):
    """Обработка пагинации."""
    page = int(callback.data.split(":")[1])
    data = await state.get_data()
    items = data.get('items', [])
    brewery_filter = data.get('brewery_filter', None)
    list_message_id = data.get('list_message_id')
    
    await state.update_data(current_page=page)
    await show_items_page(callback.message, items, page, brewery_filter=brewery_filter, edit_message_id=list_message_id, state=state)
    await callback.answer()


@router.callback_query(F.data.startswith("item_info:"))
async def handle_item_info(callback: CallbackQuery):
    """Показать информацию о позиции (заглушка)."""
    await callback.answer("Введите номер позиции в чат для выбора количества")


@router.callback_query(F.data == "page_info")
async def handle_page_info(callback: CallbackQuery):
    """Обработка нажатия на индикатор страницы."""
    await callback.answer("Используйте кнопки для навигации")


@router.callback_query(F.data == "finish_order")
async def handle_finish_callback(callback: CallbackQuery, state: FSMContext):
    """Завершить заказ через callback."""
    await callback.answer()
    await finish_order(callback.message, state)


async def show_cart_message(message: Message, items: List[Dict], state: FSMContext):
    """Показать корзину (вспомогательная функция)."""
    # Получаем выбранные позиции
    selected_items = [(i+1, item) for i, item in enumerate(items) if (item.get('заказ') or 0) > 0]
    
    if not selected_items:
        await message.answer("Корзина пуста\n\nВыберите позиции для заказа.")
        await state.set_state(QuickOrderStates.viewing_page)
        return
    
    # Формируем текст корзины
    text = f"**ВАША КОРЗИНА** ({len(selected_items)} позиций)\n\n"
    
    # Группируем по пивоварням
    breweries = {}
    total_sum = 0.0
    
    for idx, item in selected_items:
        brewery = item.get('пивоварня', 'Без пивоварни')
        if brewery not in breweries:
            breweries[brewery] = []
        
        qty = item.get('заказ') or 0
        price_str = str(item.get('цена', '')).replace('руб.', '').strip()
        try:
            price = float(price_str)
            sum_price = price * qty
            total_sum += sum_price
        except:
            sum_price = None
        
        breweries[brewery].append((idx, item, qty, sum_price))
    
    # Выводим по пивоварням
    for brewery, brewery_items in breweries.items():
        text += f"**{brewery}**\n"
        for idx, item, qty, sum_price in brewery_items:
            name = item['название']
            if len(name) > 35:
                name = name[:35] + "..."
            
            volume = item.get('объем', '')
            text += f"`{idx:3d}`. {name}\n"
            text += f"      {volume} x {qty} шт"
            
            if sum_price:
                text += f" = {sum_price:.0f}₽"
            text += "\n"
        text += "\n"
    
    if total_sum > 0:
        text += f"**Итого: {total_sum:.0f} руб.**\n"
    
    text += "\n**Редактирование:**\n"
    text += "`номер:новое_кол-во` - изменить\n"
    text += "`номер:0` - удалить из корзины\n"
    text += "Пример: `5:15` - изменить поз. 5 на 15 шт"
    
    # Кнопки
    keyboard_builder = InlineKeyboardBuilder()
    keyboard_builder.row(
        InlineKeyboardButton(text="< К списку", callback_data="back_to_list"),
        InlineKeyboardButton(text="Очистить", callback_data="clear_cart")
    )
    keyboard_builder.row(
        InlineKeyboardButton(text="Завершить заказ", callback_data="finish_order")
    )
    
    await message.answer(text, parse_mode="Markdown", reply_markup=keyboard_builder.as_markup())
    await state.set_state(QuickOrderStates.viewing_cart)


@router.callback_query(F.data == "show_cart")
async def handle_show_cart(callback: CallbackQuery, state: FSMContext):
    """Показать корзину."""
    await callback.answer()
    data = await state.get_data()
    items = data.get('items', [])
    
    await show_cart_message(callback.message, items, state)


@router.callback_query(F.data == "back_to_list")
async def handle_back_to_list(callback: CallbackQuery, state: FSMContext):
    """Вернуться к списку позиций."""
    await callback.answer()
    data = await state.get_data()
    items = data.get('items', [])
    current_page = data.get('current_page', 0)
    brewery_filter = data.get('brewery_filter', None)
    list_message_id = data.get('list_message_id')
    
    await show_items_page(callback.message, items, current_page, brewery_filter=brewery_filter, edit_message_id=list_message_id, state=state)
    await state.set_state(QuickOrderStates.viewing_page)


@router.callback_query(F.data == "clear_cart")
async def handle_clear_cart(callback: CallbackQuery, state: FSMContext):
    """Очистить корзину."""
    data = await state.get_data()
    items = data.get('items', [])
    
    # Очищаем все заказы
    for item in items:
        item['заказ'] = 0
    
    await state.update_data(items=items)
    await callback.answer("Корзина очищена")
    
    current_page = data.get('current_page', 0)
    brewery_filter = data.get('brewery_filter', None)
    list_message_id = data.get('list_message_id')
    await show_items_page(callback.message, items, current_page, brewery_filter=brewery_filter, edit_message_id=list_message_id, state=state)
    await state.set_state(QuickOrderStates.viewing_page)


@router.callback_query(F.data == "show_breweries")
async def handle_show_breweries(callback: CallbackQuery, state: FSMContext):
    """Показать список пивоварен для фильтрации."""
    await callback.answer()
    data = await state.get_data()
    items = data.get('items', [])
    
    # Получаем уникальные пивоварни с подсчетом позиций
    breweries_count = {}
    for item in items:
        brewery = item.get('пивоварня', 'Без пивоварни')
        breweries_count[brewery] = breweries_count.get(brewery, 0) + 1
    
    # Формируем текст
    text = "**ФИЛЬТР ПО ПИВОВАРНЯМ**\n\n"
    text += "Выберите пивоварню для фильтрации:\n\n"
    
    # Кнопки пивоварен
    builder = InlineKeyboardBuilder()
    for brewery, count in sorted(breweries_count.items()):
        builder.row(InlineKeyboardButton(
            text=f"{brewery} ({count})",
            callback_data=f"filter_brewery:{brewery}"
        ))
    
    builder.row(InlineKeyboardButton(
        text="< Назад",
        callback_data="back_to_list"
    ))
    
    await callback.message.answer(text, parse_mode="Markdown", reply_markup=builder.as_markup())


@router.callback_query(F.data.startswith("filter_brewery:"))
async def handle_filter_brewery(callback: CallbackQuery, state: FSMContext):
    """Применить фильтр по пивоварне."""
    brewery = callback.data.split(":", 1)[1]
    await callback.answer(f"Фильтр: {brewery}")
    
    data = await state.get_data()
    items = data.get('items', [])
    list_message_id = data.get('list_message_id')
    
    # Сохраняем фильтр
    await state.update_data(brewery_filter=brewery, current_page=0)
    
    sent_msg = await show_items_page(callback.message, items, 0, brewery_filter=brewery, edit_message_id=list_message_id, state=state)
    if sent_msg:
        await state.update_data(list_message_id=sent_msg.message_id)
    await state.set_state(QuickOrderStates.viewing_page)


@router.callback_query(F.data == "clear_filter")
async def handle_clear_filter(callback: CallbackQuery, state: FSMContext):
    """Сбросить фильтр по пивоварне."""
    await callback.answer("Фильтр сброшен")
    
    data = await state.get_data()
    items = data.get('items', [])
    list_message_id = data.get('list_message_id')
    
    await state.update_data(brewery_filter=None, current_page=0)
    sent_msg = await show_items_page(callback.message, items, 0, edit_message_id=list_message_id, state=state)
    if sent_msg:
        await state.update_data(list_message_id=sent_msg.message_id)
    await state.set_state(QuickOrderStates.viewing_page)


@router.callback_query(F.data == "start_search")
async def handle_start_search(callback: CallbackQuery, state: FSMContext):
    """Начать поиск."""
    await callback.answer()
    
    text = "**ПОИСК**\n\n"
    text += "Введите название пива для поиска:\n"
    text += "Например: `ghost`, `helles`, `ipa`\n\n"
    text += "Отправьте `/cancel` для отмены"
    
    builder = InlineKeyboardBuilder()
    builder.row(InlineKeyboardButton(text="Отмена", callback_data="back_to_list"))
    
    await callback.message.answer(text, parse_mode="Markdown", reply_markup=builder.as_markup())
    await state.set_state(QuickOrderStates.searching)


@router.message(QuickOrderStates.searching)
async def process_search(message: Message, state: FSMContext):
    """Обработать поисковый запрос."""
    if message.text == "/cancel":
        data = await state.get_data()
        items = data.get('items', [])
        current_page = data.get('current_page', 0)
        brewery_filter = data.get('brewery_filter', None)
        sent_msg = await show_items_page(message, items, current_page, brewery_filter=brewery_filter, state=state)
        if sent_msg:
            await state.update_data(list_message_id=sent_msg.message_id)
        await state.set_state(QuickOrderStates.viewing_page)
        return
    
    search_query = message.text.lower().strip()
    data = await state.get_data()
    items = data.get('items', [])
    
    # Ищем совпадения
    found_items = []
    for i, item in enumerate(items):
        name = item.get('название', '').lower()
        brewery = item.get('пивоварня', '').lower()
        style = item.get('стиль', '').lower()
        
        if search_query in name or search_query in brewery or search_query in style:
            found_items.append((i+1, item))
    
    if not found_items:
        await message.answer(f"Ничего не найдено по запросу: `{search_query}`\n\nПопробуйте другой запрос.", parse_mode="Markdown")
        return
    
    # Показываем результаты
    text = f"**РЕЗУЛЬТАТЫ ПОИСКА:** `{search_query}`\n\n"
    text += f"Найдено: {len(found_items)} позиций\n\n"
    
    # Показываем первые 20 результатов
    for idx, item in found_items[:20]:
        name = item['название']
        if len(name) > 35:
            name = name[:35] + "..."
        
        qty = item.get('заказ') or 0
        checkbox = "✓" if qty > 0 else " "
        
        brewery = item.get('пивоварня', '')
        volume = item.get('объем', '')
        price = item.get('цена', '')
        
        text += f"`{idx:3d}` [{checkbox}] {name}"
        if qty > 0:
            text += f" **x{qty}**"
        text += f"\n      {brewery} | {volume} | {price}\n"
    
    if len(found_items) > 20:
        text += f"\n... и еще {len(found_items) - 20} позиций"
    
    text += "\n\nВведите номера для заказа или выберите действие:"
    
    # Кнопки
    builder = InlineKeyboardBuilder()
    builder.row(
        InlineKeyboardButton(text="< К списку", callback_data="back_to_list"),
        InlineKeyboardButton(text="Новый поиск", callback_data="start_search")
    )
    builder.row(InlineKeyboardButton(text="Завершить заказ", callback_data="finish_order"))
    
    await message.answer(text, parse_mode="Markdown", reply_markup=builder.as_markup())
    await state.set_state(QuickOrderStates.viewing_page)


@router.message(QuickOrderStates.viewing_cart)
async def process_cart_edit(message: Message, state: FSMContext):
    """Обработать редактирование из корзины."""
    data = await state.get_data()
    items = data.get('items', [])
    
    # Парсим изменения
    text = message.text.replace(',', ' ')
    changes = {}
    
    for part in text.split():
        if ':' in part:
            try:
                num_str, qty_str = part.split(':', 1)
                num = int(num_str)
                qty = int(qty_str)
                if 1 <= num <= len(items):
                    changes[num] = qty
            except ValueError:
                continue
    
    if not changes:
        await message.answer("Не распознаны изменения.\n\nФормат: `номер:количество`", parse_mode="Markdown")
        return
    
    # Применяем изменения
    updated_items = []
    for num, qty in changes.items():
        items[num - 1]['заказ'] = qty
        updated_items.append((num, qty, items[num-1]['название']))
    
    await state.update_data(items=items)
    
    # Сообщаем об изменениях
    change_text = "Обновлено:\n"
    for num, qty, name in updated_items[:5]:
        short_name = name[:25] + "..." if len(name) > 25 else name
        if qty == 0:
            change_text += f"  Поз. {num}: удалено\n"
        else:
            change_text += f"  Поз. {num}: {short_name} -> {qty} шт\n"
    
    await message.answer(change_text)
    
    # Показываем обновленную корзину
    await show_cart_message(message, items, state)


@router.message(QuickOrderStates.viewing_page)
async def process_quantities(message: Message, state: FSMContext):
    """
    Обработать указанные количества.
    """
    if message.text == "/finish":
        await finish_order(message, state)
        return
    
    data = await state.get_data()
    items = data.get('items', [])
    
    # Парсим введенные номера с количеством
    text = message.text.replace(',', ' ')
    quantities = {}  # {номер: количество}
    
    for part in text.split():
        if ':' in part:
            # Формат "номер:количество"
            try:
                num_str, qty_str = part.split(':', 1)
                num = int(num_str)
                qty = int(qty_str)
                if 1 <= num <= len(items):
                    quantities[num] = qty
            except ValueError:
                continue
        else:
            # Просто номер (количество = 1)
            try:
                num = int(part)
                if 1 <= num <= len(items):
                    quantities[num] = 1
            except ValueError:
                continue
    
    if not quantities:
        await message.answer(
            "Не распознаны позиции.\n\n"
            "Формат:\n"
            "`1 3 5` - выбрать 1, 3, 5 (по 1 шт)\n"
            "`1:10 3:5` - позиция 1 (10 шт), 3 (5 шт)",
            parse_mode="Markdown"
        )
        return
    
    # Обновляем количество в items
    for num, qty in quantities.items():
        if 1 <= num <= len(items):
            items[num - 1]['заказ'] = qty  # Индексация с 0
    
    # Сохраняем обновленные данные
    await state.update_data(items=items)
    
    # Показываем что добавлено
    selected_count = len(quantities)
    total_qty = sum(quantities.values())
    
    # Формируем список добавленных позиций
    added_list = []
    for num, qty in quantities.items():
        if 1 <= num <= len(items):
            item = items[num - 1]
            name = item.get('название', 'Без названия')
            volume = item.get('объем', '')
            if len(name) > 30:
                name = name[:30] + "..."
            added_list.append(f"• {name} ({volume}) x{qty}")
    
    response_text = f"Добавлено: {selected_count} позиций ({total_qty} шт)\n\n"
    if added_list:
        response_text += "\n".join(added_list[:5])  # Показываем первые 5
        if len(added_list) > 5:
            response_text += f"\n... и еще {len(added_list) - 5} позиций"
    
    response_text += "\n\nПродолжайте выбирать позиции или нажмите 'Завершить заказ'."
    
    await message.answer(response_text, parse_mode="Markdown")


async def finish_order(message: Message, state: FSMContext):
    """Завершить заказ и сгенерировать Excel."""
    data = await state.get_data()
    file_path = data.get('file_path')
    filename = data.get('filename')
    items = data.get('items', [])
    
    # Проверяем есть ли выбранные позиции
    selected_items = [item for item in items if (item.get('заказ') or 0) > 0]
    if not selected_items:
        await message.answer("Вы не выбрали ни одной позиции для заказа.\n\nСоздается пустой файл.")
    
    await message.answer("Генерация Excel файла...")
    
    # Генерируем Excel с заполненной колонкой "Заказ"
    excel_bytes = generate_excel_with_order(items, file_path)
    
    # Формируем имя файла: Число.месяц.год-название поставщика.расширение
    now = datetime.now()
    date_str = f"{now.day:02d}.{now.month:02d}.{now.year}"
    
    # Получаем название поставщика из имени файла
    original_name = Path(filename).stem  # Без расширения
    
    # Убираем префиксы с датами (например "08_10 ", "24_09 ")
    name_without_date = re.sub(r'^\d{1,2}[_-]\d{1,2}\s+', '', original_name)
    
    # Берем часть до первой точки
    supplier_name = name_without_date.split('.')[0].strip()
    
    # Убираем лишние слова и суффиксы (case-insensitive, сначала длинные фразы)
    words_to_remove = [
        'актуальный прайс', 'для бота', 'pricelist', 'price list',
        'актуальный', 'прайс', 'price', 'excel', 'список', 'list'
    ]
    supplier_name_lower = supplier_name.lower()
    for word in words_to_remove:
        if word.lower() in supplier_name_lower:
            # Case-insensitive замена
            supplier_name = re.sub(re.escape(word), '', supplier_name, flags=re.IGNORECASE)
            supplier_name_lower = supplier_name.lower()
    supplier_name = supplier_name.strip()
    
    # Убираем подчеркивания и лишние пробелы в конце/начале
    supplier_name = re.sub(r'[_\s]+$', '', supplier_name)  # Убираем _ и пробелы в конце
    supplier_name = re.sub(r'^[_\s]+', '', supplier_name)  # Убираем _ и пробелы в начале
    supplier_name = supplier_name.strip()
    
    # Приводим к нижнему регистру
    supplier_name = supplier_name.lower()
    
    # Заменяем оставшиеся подчеркивания на пробелы для красоты
    supplier_name = supplier_name.replace('_', ' ')
    
    file_ext = Path(filename).suffix
    output_filename = f"{date_str}-{supplier_name}{file_ext}"
    
    # Сохраняем в базу данных
    async with async_session_maker() as session:
        user = await get_or_create_user(
            session,
            telegram_id=message.from_user.id,
            username=message.from_user.username
        )
        
        await create_quick_order(
            session,
            user_id=user.id,
            filename=filename,
            original_data=json.dumps(items, ensure_ascii=False),
            order_data=json.dumps(selected_items, ensure_ascii=False)
        )
    
    # Отправляем файл
    file_obj = BufferedInputFile(
        excel_bytes.getvalue(),
        filename=output_filename
    )
    
    await message.answer("Заказ сформирован!")
    await message.answer_document(file_obj, caption=f"Заказ от {date_str}")
    
    # Очищаем состояние
    await state.clear()


def generate_excel_with_order(items: List[Dict], original_file_path: str) -> BytesIO:
    """
    Сгенерировать Excel файл с заполненной колонкой "Заказ" в оригинальных листах.
    Сохраняет ВСЁ форматирование оригинала.
    
    Args:
        items: Список позиций с количеством заказа
        original_file_path: Путь к оригинальному файлу
        
    Returns:
        BytesIO: Excel файл в памяти
    """
    # Создаем маппинг: (название, sheet_index) -> количество
    name_to_qty = {}
    for item in items:
        name = item.get('название')
        qty = item.get('заказ')
        sheet_idx = item.get('_sheet_index', 0)
        if name and qty and qty > 0:
            name_to_qty[(name, sheet_idx)] = qty
    
    # Если файл в формате .xls, конвертируем в .xlsx
    file_to_open = original_file_path
    temp_xlsx_path = None
    
    if original_file_path.endswith('.xls'):
        # Конвертируем .xls в .xlsx через pandas
        temp_xlsx_path = original_file_path.replace('.xls', '_temp.xlsx')
        xls = pd.ExcelFile(original_file_path, engine='xlrd')
        with pd.ExcelWriter(temp_xlsx_path, engine='openpyxl') as writer:
            for sheet_name in xls.sheet_names:
                df = pd.read_excel(xls, sheet_name=sheet_name)
                df.to_excel(writer, sheet_name=sheet_name, index=False)
        file_to_open = temp_xlsx_path
    
    # Открываем файл через openpyxl (сохраняет форматирование для .xlsx)
    wb = load_workbook(file_to_open)
    
    # Обрабатываем каждый лист
    for sheet_idx, sheet_name in enumerate(wb.sheetnames):
        ws = wb[sheet_name]
        
        # Шаг 1: Найти строку с заголовками
        # Ищем строку где есть ключевые заголовки (Цена, Название и т.д.)
        header_row = None
        best_score = 0
        
        for row_idx in range(1, min(20, ws.max_row + 1)):
            score = 0
            for col_idx in range(1, ws.max_column + 1):
                cell_value = str(ws.cell(row=row_idx, column=col_idx).value or "").lower().strip()
                # Баллы за ключевые заголовки
                if cell_value in ['название', 'наименование', 'name', 'продукт']:
                    score += 10
                if cell_value in ['цена', 'price', 'стоимость']:
                    score += 10
                if cell_value in ['стиль', 'style']:
                    score += 5
                if 'пивоварн' in cell_value or 'brewery' in cell_value:
                    score += 5
            
            # Строка с минимум 20 баллами (название + цена) - это заголовки
            if score >= 20 and score > best_score:
                best_score = score
                header_row = row_idx
        
        # Если не нашли - используем строку 1
        if header_row is None:
            header_row = 1
        
        # Шаг 2: Ищем колонку "Заказ" в найденной строке заголовков
        order_col_idx = None
        
        for col_idx in range(1, ws.max_column + 1):
            cell_value = str(ws.cell(row=header_row, column=col_idx).value or "").lower().strip()
            if 'заказ' in cell_value or 'order' in cell_value:
                order_col_idx = col_idx
                break
        
        # Если колонка "Заказ" не найдена - создаем её в строке заголовков
        if order_col_idx is None:
            order_col_idx = ws.max_column + 1
            ws.cell(row=header_row, column=order_col_idx, value="Заказ")
        
        # Заполняем колонку "Заказ" значениями только для выбранных позиций
        for item in items:
            if item.get('_sheet_index') == sheet_idx:
                qty = item.get('заказ')
                row_idx = item.get('_row_index')
                
                # Проверяем корректность индексов и количество
                if qty and qty > 0 and isinstance(row_idx, int):
                    if 1 <= row_idx <= ws.max_row:
                        cell = ws.cell(row=row_idx, column=order_col_idx)
                        cell.value = qty
    
    # Сохраняем в BytesIO
    output = BytesIO()
    wb.save(output)
    output.seek(0)
    
    # Удаляем временный файл если он был создан
    if temp_xlsx_path and os.path.exists(temp_xlsx_path):
        try:
            os.remove(temp_xlsx_path)
        except Exception:
            pass  # Игнорируем ошибки удаления
    
    return output

